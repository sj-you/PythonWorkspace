from openpyxl import load_workbook
wb = load_workbook("semple.xlsx")
ws = wb.active

ws.move_range("B1:C11", rows=0, cols=1)
ws["B1"].value = "국어"

wb.save("semple_korean.xlsx")